#input: full path to xlsx file containing article titles from scopus. Assumes that 
#titles are first column in the sheet 'Qatar_Scopus'. 
#output: csv file containing paper titles, citations, possibly free PDF link from web. 
from openpyxl import load_workbook
import sys
from scholar import ScholarQuerier, ScholarSettings, SearchScholarQuery,onecsv
from time import sleep
import csv

querier = ScholarQuerier()
settings = ScholarSettings()
querier.apply_settings(settings)
query = SearchScholarQuery()
query.set_scope(True)
alldata=[]
counter=1

xlsxfile=sys.argv[1]
wb = load_workbook(xlsxfile, use_iterators=True)
print wb.get_sheet_names()
ws = wb.get_sheet_by_name('Qatar_Scopus')

for row in ws.iter_rows(row_offset=1):
	if row[0].value is not None:
		temp=[]
		title=row[0].value.encode("utf-8")
		query.set_phrase(title)
		query.set_num_page_results(1)
		querier.send_query(query)
 	   	x=onecsv(querier)
	  	if (x!=None):
   			year=x.split("|")[2]
   			numcit=x.split("|")[3]
   			weburl=x.split("|")[1]
   		else:
			year=None;numcit=None;weburl=None;	
   		temp.append(title),temp.append(numcit);temp.append(year);temp.append(weburl);
   		alldata.append(temp)
  		print "Title: ",title,"query number: ",counter,"No. citations: ",numcit
		print "sleeeping for 5 seconds"
		sleep(5)
		counter=counter+1

with open('scopus-2010-2014-cit.csv', 'wb') as fp:
    a = csv.writer(fp, delimiter=',')
    a.writerows(alldata)
				

